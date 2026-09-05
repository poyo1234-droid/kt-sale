# -*- coding: utf-8 -*-
"""キントーン「セール返答表」(k/60) の添付ファイルを一括ダウンロードする。

精算書の `_スクリプト/download_attachments-2607.py` と同じ作りで、
参照するアプリと保存名の付け方だけ本件に合わせてある。

保存名は添付名そのまま（先頭がメーカーNo.）なので、落としたフォルダを
そのまま `キントーン突合.py --kintone <フォルダ>` に渡せる。

  python キントーン添付DL.py --id 1736-1938
  python キントーン添付DL.py --セール設定 3 --以降 2026-09-05T09:40
  python キントーン添付DL.py --クエリ "$id >= 1736 and $id <= 1938"

対象は必ずどれかで絞る。何も指定しないと止まる（--全件 を書いたときだけ全件）。

前提: pip install requests openpyxl
"""

import argparse
import csv
import os
import sys
import time

import openpyxl
import requests

# --- パス（このファイルの位置から解決する）--------------------------------
HERE   = os.path.dirname(os.path.abspath(__file__))                 # 検証スクリプト
REPO   = os.path.dirname(HERE)                                      # エクセル開発
BASE   = os.path.dirname(REPO)                                      # 251016-SALEデータ一式-今野様
SERVER = os.path.dirname(BASE)                                      # 250829サーバー更新

CONFIG_PATH = os.path.join(
    SERVER, "260318ー登録エクセルkintone化", "納品書作成", "config", "kintone_config.xlsx")

OUTPUT_DIR = os.path.join(
    BASE, "260904-kintone動作確認", "セール情報設定リスト 二期間分",
    "セール情報設定リスト202608089～202608219", "kintone出力")

# --- kintone_config.xlsx のセル位置 ---------------------------------------
# A65 = 「セール返答表」/ C65 = APIトークン / C66 = アプリID / C3 = サブドメイン
CELL_LABEL     = "A65"
CELL_TOKEN     = "C65"
CELL_APP_ID    = "C66"
CELL_SUBDOMAIN = "C3"
EXPECT_LABEL   = "セール返答表"

FILE_FIELD = "セール価格設定リスト"          # 添付ファイルのフィールドコード
FIELDS = ["$id", "メーカーコード", "メーカー名", "ブランド名", "親カテゴリーコード",
          "セール設定レコード番号", "セール開始日", "セール終了日",
          "ブランド連絡先担当者名", "ブランド連絡先メールアドレス", "FA様担当者",
          "作成日時", FILE_FIELD]
SLEEP_SEC = 0.1


def load_config(path):
    """設定ブックから接続情報を読む。行がずれていたら止める。"""
    if not os.path.exists(path):
        sys.exit("設定ファイルが見つかりません: %s\n--config でパスを指定してください。" % path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["設定"] if "設定" in wb.sheetnames else wb.active

    label = str(ws[CELL_LABEL].value or "").strip()
    if label != EXPECT_LABEL:
        sys.exit(
            "設定ファイルの %s が「%s」ではありません（実際: 「%s」）。\n"
            "行が増減した可能性があります。CELL_LABEL / CELL_TOKEN / CELL_APP_ID を直してください。"
            % (CELL_LABEL, EXPECT_LABEL, label))

    subdomain = str(ws[CELL_SUBDOMAIN].value or "").strip()
    token     = str(ws[CELL_TOKEN].value or "").strip()
    app_id    = ws[CELL_APP_ID].value
    if not subdomain or not token or not app_id:
        sys.exit("設定ファイルの %s / %s / %s のいずれかが空です。" % (CELL_SUBDOMAIN, CELL_TOKEN, CELL_APP_ID))
    return subdomain, token, int(app_id)


def build_query(args):
    """コマンドラインから kintone のクエリを組み立てる。"""
    conds = []
    if args.クエリ:
        conds.append("(%s)" % args.クエリ)
    if args.id:
        s = args.id.replace("〜", "-").replace("～", "-")
        if "-" in s:
            a, b = [t.strip() for t in s.split("-", 1)]
            conds.append('$id >= %d and $id <= %d' % (int(a), int(b)))
        else:
            conds.append('$id = %d' % int(s))
    if args.セール設定 is not None:
        conds.append('セール設定レコード番号 = %d' % args.セール設定)
    if args.以降:
        t = args.以降
        if len(t) == 16:          # 2026-09-05T09:40 → 秒とZを補う
            t += ":00Z"
        elif not t.endswith("Z"):
            t += "Z"
        conds.append('作成日時 >= "%s"' % t)

    if not conds:
        if not args.全件:
            sys.exit("対象が指定されていません。--id / --セール設定 / --以降 / --クエリ の"
                     "いずれかで絞るか、全部落とすなら --全件 を付けてください。")
        return ""
    return " and ".join(conds)


def get_records(base_url, headers, app_id, query):
    """$id を進めながら全件取る（offset を使わない = 5,000件の壁が無い）。"""
    records = []
    last_id = 0
    while True:
        q = "%s and $id > %d" % (query, last_id) if query else "$id > %d" % last_id
        q += " order by $id asc limit 500"
        res = requests.get(base_url + "/k/v1/records.json", headers=headers,
                           params={"app": app_id, "query": q, "fields": FIELDS})
        if res.status_code != 200:
            sys.exit("レコード取得に失敗しました [%d]\n%s" % (res.status_code, res.text[:500]))
        batch = res.json()["records"]
        if not batch:
            break
        records.extend(batch)
        last_id = int(batch[-1]["$id"]["value"])
        print("  取得済み: %d 件" % len(records))
        if len(batch) < 500:
            break
        time.sleep(SLEEP_SEC)
    return records


def download_file(base_url, headers, file_key, save_path):
    res = requests.get(base_url + "/k/v1/file.json", headers=headers,
                       params={"fileKey": file_key}, stream=True)
    res.raise_for_status()
    tmp = save_path + ".part"
    with open(tmp, "wb") as f:
        for chunk in res.iter_content(chunk_size=8192):
            f.write(chunk)
    os.replace(tmp, save_path)


def val(rec, code):
    v = rec.get(code, {}).get("value")
    if isinstance(v, list):                       # FA様担当者（ユーザー選択）
        return ",".join(x.get("name", "") for x in v)
    return "" if v is None else str(v)


def main():
    p = argparse.ArgumentParser(description="k/60 セール返答表の添付を一括ダウンロードする")
    p.add_argument("--id", help='レコードIDの範囲。例: 1736-1938 / 単体なら 1736')
    p.add_argument("--セール設定", type=int, help="セール設定レコード番号で絞る")
    p.add_argument("--以降", help='作成日時の下限。例: 2026-09-05T09:40（UTC）')
    p.add_argument("--クエリ", help="kintone のクエリをそのまま書く")
    p.add_argument("--全件", action="store_true", help="絞り込みなしで全件落とす")
    p.add_argument("--出力先", default=OUTPUT_DIR, help="保存先フォルダ")
    p.add_argument("--config", default=CONFIG_PATH, help="kintone_config.xlsx のパス")
    p.add_argument("--上書き", action="store_true", help="同名ファイルがあっても落とし直す")
    args = p.parse_args()

    subdomain, token, app_id = load_config(args.config)
    base_url = "https://%s.cybozu.com" % subdomain
    headers  = {"X-Cybozu-API-Token": token}
    query    = build_query(args)

    out = args.出力先
    os.makedirs(out, exist_ok=True)

    print("接続先 : %s / アプリ %d" % (base_url, app_id))
    print("対象   : %s" % (query if query else "全件"))
    print("保存先 : %s\n" % out)

    print("レコード取得中...")
    records = get_records(base_url, headers, app_id, query)
    print("対象レコード数: %d 件\n" % len(records))
    if not records:
        return

    ok = skip = err = nofile = 0
    rows = []
    used = {}
    for rec in records:
        rid   = rec["$id"]["value"]
        files = rec.get(FILE_FIELD, {}).get("value", [])
        if not files:
            nofile += 1
            print("  [なし] レコード %s %s" % (rid, val(rec, "メーカー名")))
        for fi in files:
            name = fi["name"]
            # 同名が別レコードから来たときだけレコードIDを足す（通常は添付名のまま）
            if name in used:
                stem, ext = os.path.splitext(name)
                name = "%s_r%s%s" % (stem, rid, ext)
            used[name] = rid
            path = os.path.join(out, name)

            rows.append({
                "レコードID": rid,
                "メーカーコード": val(rec, "メーカーコード"),
                "メーカー名": val(rec, "メーカー名"),
                "ブランド名": val(rec, "ブランド名"),
                "親カテゴリーコード": val(rec, "親カテゴリーコード"),
                "セール設定レコード番号": val(rec, "セール設定レコード番号"),
                "セール開始日": val(rec, "セール開始日"),
                "セール終了日": val(rec, "セール終了日"),
                "ブランド連絡先担当者名": val(rec, "ブランド連絡先担当者名"),
                "ブランド連絡先メールアドレス": val(rec, "ブランド連絡先メールアドレス"),
                "FA様担当者": val(rec, "FA様担当者"),
                "作成日時": val(rec, "作成日時"),
                "ファイル名": name,
            })

            if os.path.exists(path) and not args.上書き:
                print("  [SKIP] %s" % name)
                skip += 1
                continue
            try:
                download_file(base_url, headers, fi["fileKey"], path)
                print("  [OK]   %s" % name)
                ok += 1
                time.sleep(SLEEP_SEC)
            except Exception as e:
                print("  [ERR]  %s -> %s" % (name, e))
                err += 1

    # 宛先3列を含む一覧。突合では見えない「宛先が入っているか」をここで見る。
    if rows:
        csv_path = os.path.join(out, "_DL一覧.csv")
        with open(csv_path, "w", encoding="cp932", errors="replace", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        print("\n一覧: %s" % csv_path)

    blank = sum(1 for r in rows if not r["ブランド連絡先メールアドレス"] and not r["FA様担当者"])
    print("\n完了: %d 件DL / %d 件スキップ / %d 件エラー / 添付なし %d 件" % (ok, skip, err, nofile))
    print("宛先（メール・FA様担当者）が両方とも空: %d 件" % blank)


if __name__ == "__main__":
    main()
