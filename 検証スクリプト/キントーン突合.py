"""キントーンが出力したセール価格設定リストを、旧システム（SQL ベースの基幹システム）の出力と突き合わせる。

  python キントーン突合.py                        # 既定のフォルダ同士を全メーカー突合
  python キントーン突合.py --メーカー 2 4          # メーカーNo. を指定して絞る
  python キントーン突合.py --kintone <dir> --旧 <dir> --out <path>

キントーン側は 23列（A:親カテゴリ 〜 W:直近30日お気に入り登録数）、
旧システム側は 45列（A:メーカーNo. 〜 AJ:精算掛率）と列構成が異なるため、
下の ATTR / PERF / ONLY_KYUU で対応付けている。

比較は 2 種類に分けている。どちらも本来は一致すべきで、分けているのは原因の性質が違うため。

  属性（ATTR）… 品名・カテゴリ・価格など、抽出元に依らない項目。
                 不一致は移行ロジックの問題を指す
  実績（PERF）… 在庫・販売数・セール実績など、抽出元データに依存する項目。
                 不一致は抽出条件（どの列を採るか、集計期間の取り方）の問題を指す

  ※ 同じ回のリスト同士を突き合わせる前提。実行日が違っても実績値は一致するはず。

突合キーは ブランド品番（キントーン C列 / 旧 E列）。
結果は 6 シートの Excel（既定では キントーン側フォルダ）に出力する。
"""
import os, sys, glob, re, argparse, datetime, collections, zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl

from _共通 import KINTONE_LIST, KYUU_LIST

# --- 列マッピング（1始まり）------------------------------------------------
# (表示名, キントーン列, 旧システム列)

#: 抽出元に依らない項目。不一致＝移行ロジックの問題
ATTR = [
    ("親カテゴリ",           1,  3),
    ("ブランド(子カテゴリ)", 2,  4),
    ("主性別",               4,  6),
    ("商品タイプ(親)",       5,  7),
    ("商品タイプ(子)",       6,  8),
    ("品名",                 7,  9),
    ("プロパー価格(元上代)", 8, 11),
    ("セール価格(税抜)",    18, 18),
    ("セール価格(税込)",    17, 19),
    ("オフ率",              16, 20),
]

#: 抽出元データに依存する項目。不一致＝抽出条件の問題
PERF = [
    ("販売開始日",           19, 24),
    ("在庫",                 20, 26),
    ("直近30日販売数",       21, 27),
    ("直近7日販売数",        22, 28),
    ("お気に入り登録数",     23, 29),
    ("受注数",                9, 30),
    ("売上額",               10, 31),
    ("直近セール価格(税抜)", 11, 32),
    ("直近セールオフ率",     12, 33),
]

#: 旧システムにあってキントーンに無い項目
ONLY_KYUU = ["メーカーNo.(A列)", "メーカー名(B列)", "SKU数(J列)", "MAX割引額(L列)",
             "MAX割引率(M列)", "直近割引率(N列)", "値引き金額 入力(P列)", "販売タイプ(V列)",
             "初回販売開始日(W列)", "販売可能数(Y列)", "セール前在庫数(AH列)", "精算掛率(AJ列)"]

#: キントーンにあって旧システムに無い項目
ONLY_KINTONE = ["直近セール価格(税込み)(M列)"]

KINTONE_NAME = {1: "親カテゴリ", 2: "子カテゴリ", 3: "ブランド品番", 4: "主性別", 5: "(親)",
                6: "(子)", 7: "品名", 8: "プロパー価格（税抜）", 9: "受注数", 10: "売上額",
                11: "直近セール価格（税抜き）", 12: "直近セールオフ率", 13: "直近セール価格（税込み）",
                14: "オフ率から設定(↓)", 15: "税込価格から設定(↓)", 16: "オフ率",
                17: "セール価格（税込み）", 18: "セール価格（税抜き）", 19: "販売開始日",
                20: "在庫", 21: "直近30日販売数", 22: "直近7日販売数", 23: "直近30日お気に入り登録数"}
KYUU_NAME = {3: "親カテゴリー", 4: "ブランド", 5: "ブランド品番", 6: "主性別", 7: "(親)", 8: "(子)",
             9: "品名", 11: "元上代", 15: "オフ率 入力(↓)優先", 17: "税込価格から設定 入力(↓)",
             18: "税抜価格(設定価格)", 19: "税込表示", 20: "オフ率", 24: "販売開始日", 26: "在庫",
             27: "直近30日販売数", 28: "直近7日販売数", 29: "お気に入り登録数", 30: "受注数",
             31: "売上額", 32: "SALE価格", 33: "SALEオフ率"}

KINTONE_KEYCOL, KINTONE_NCOL = 3, 23
KYUU_KEYCOL,    KYUU_NCOL    = 5, 45
KINTONE_STOCK, KINTONE_NAME_COL = 20, 7      # 在庫 / 品名（1始まり）
DATA_ROW = 3                                  # 1行目=グループ見出し 2行目=項目名

# 健全性チェックで使う列（1始まり）
KINTONE_受注数, KINTONE_売上額, KINTONE_プロパー = 9, 10, 8
KYUU_受注数,    KYUU_売上額,    KYUU_プロパー    = 30, 31, 11
KINTONE_最終列 = 23                            # A〜W


# --- 読み込み --------------------------------------------------------------

def メーカーNo(filename):
    """ファイル名の先頭の数字をメーカーNo. として取り出す。

    キントーン  '2_株式会社 ○○様_セール価格設定_…xlsx'      → '2'
    旧システム  '00002-株式会社 ○○様 セール価格設定リスト…' → '2'
    """
    m = re.match(r"^(\d+)", os.path.basename(filename))
    return str(int(m.group(1))) if m else None


def 一覧(folder):
    """フォルダ内の xlsx を メーカーNo. → [パス] にまとめる。

    ~$ の一時ファイルと、このスクリプト自身が出した突合結果は無視する。
    """
    out = collections.defaultdict(list)
    for p in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        base = os.path.basename(p)
        if base.startswith("~$") or "突合結果" in base:
            continue
        no = メーカーNo(p)
        if no:
            out[no].append(p)
    return out


def 読む(paths, keycol, ncol):
    """セール価格設定リストを ブランド品番 → 行（値のリスト）にして返す。

    旧システムは 1メーカーが子カテゴリ単位で複数ファイルに分かれることがあるため、
    複数パスを受け取って 1つの dict にまとめる。
    """
    rows, dup = {}, []
    for path in paths:
        ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]
        for r in range(DATA_ROW, ws.max_row + 1):
            key = ws.cell(r, keycol).value
            if key in (None, ""):
                continue
            key = str(key).strip()
            vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
            if key in rows:
                dup.append((key, os.path.basename(path)))
                continue
            rows[key] = vals
    return rows, dup


def 正規化(v):
    """型の違い（日付・float・全角空白）を吸収して文字列にする。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip().replace("　", " ").replace("  ", " ")


def 在庫あり(row):
    v = row[KINTONE_STOCK - 1]
    return isinstance(v, (int, float)) and v > 0


# --- 突合 ------------------------------------------------------------------

def 突合(kintone_rows, kyuu_rows):
    """1メーカー分を突き合わせ、結果を dict で返す。"""
    common = sorted(set(kintone_rows) & set(kyuu_rows))
    res = dict(
        common=common,
        kintone_only=sorted(set(kintone_rows) - set(kyuu_rows)),
        kyuu_only=sorted(set(kyuu_rows) - set(kintone_rows)),
        attr_mis=collections.Counter(), perf_mis=collections.Counter(),
        zero_pair=collections.Counter(),
        rows_attr=[], rows_perf=[],
        kintone_stock=sum(1 for v in kintone_rows.values() if 在庫あり(v)),
    )
    for key in common:
        k, o = kintone_rows[key], kyuu_rows[key]
        name = 正規化(k[KINTONE_NAME_COL - 1])
        for defs, cnt, bucket in ((ATTR, res["attr_mis"], res["rows_attr"]),
                                  (PERF, res["perf_mis"], res["rows_perf"])):
            for label, ki, oi in defs:
                a, b = 正規化(k[ki - 1]), 正規化(o[oi - 1])
                if a != b:
                    cnt[label] += 1
                    bucket.append([key, name, label, b, a])
                elif a in ("", "0"):
                    # 両方とも空欄／0 の「中身のない一致」。
                    # 一致率を水増しするので別に数える（2026-09-05 追加）
                    res["zero_pair"][label] += 1
    return res


# --- 健全性チェック（2026-09-05 追加）--------------------------------------
# 突合の「一致件数」だけを見ていると気づけない異常を検出する。
# 過去に実際に起きたものを検査項目にしてある。

def _数値(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


def _オートフィルタ(path):
    """xlsx の autoFilter ref とシートの最終行を返す。読めなければ (None, None)。"""
    try:
        with zipfile.ZipFile(path) as z:
            name = next(n for n in z.namelist() if n.startswith("xl/worksheets/sheet"))
            xml = z.read(name).decode("utf-8", errors="replace")
        m = re.search(r'<autoFilter ref="([^"]+)"', xml)
        rows = [int(x) for x in re.findall(r'<row r="(\d+)"', xml)]
        return (m.group(1) if m else None), (max(rows) if rows else None)
    except Exception:
        return None, None


def 健全性(K, O, res, kintone_paths):
    """1メーカー分の警告メッセージのリストを返す。空なら異常なし。"""
    w = []
    common = res["common"]

    # (1) 受注数・売上額が全行ゼロ
    #     2026-09-04 の出力で実際に起きた。売上の絞り込み期間が今回セール期間に
    #     なっており1行も該当しなかった。仕様書 §12-3-4
    k有 = sum(1 for v in K.values() if (_数値(v[KINTONE_受注数 - 1]) or 0) > 0)
    o有 = sum(1 for v in O.values() if (_数値(v[KYUU_受注数 - 1]) or 0) > 0)
    if k有 == 0 and o有 > 0:
        w.append(f"受注数が全行ゼロ（キントーン 0行 / 旧 {o有}行）。"
                 f"売上の絞り込み期間を疑うこと（仕様書 §12-3-4）")

    # (2) 中身のない一致（両方とも空欄／0）
    #     ゼロ同士の一致を一致と数えると、直っていないのに一致率が上がる
    if common:
        for label, n in res["zero_pair"].most_common():
            if n >= len(common) * 0.9:
                w.append(f"「{label}」は共通{len(common)}件中{n}件が両方とも空欄／0。"
                         f"一致しているのではなく、両方とも値が無いだけ")

    # (3) プロパー価格が値下げ後の疑い
    #     キントーン値 ÷ 旧値 がきれいな割引率に揃うなら、値下げ後価格を拾っている。
    #     2026-09-04 に15品番で発生。原因は在庫分析アプリの列取り違え（仕様書 §12-3-2）
    比 = collections.Counter()
    for key in common:
        a = _数値(K[key][KINTONE_プロパー - 1])
        b = _数値(O[key][KYUU_プロパー - 1])
        if a and b and a != b and 0.4 <= a / b < 1.0:
            比[round(a / b, 2)] += 1
    if 比:
        内訳 = " / ".join(f"{r}×{n}件" for r, n in 比.most_common(5))
        w.append(f"プロパー価格が旧より安い品番が{sum(比.values())}件（比率 {内訳}）。"
                 f"値下げ後の価格を拾っている疑い（仕様書 §12-3-2）")

    # (4) オートフィルタの範囲と空行
    #     テンプレートが2000行固定なので、データ行数と合わないことがある（仕様書 §9-4）
    最終データ行 = DATA_ROW + len(K) - 1
    for p in kintone_paths:
        ref, 最終行 = _オートフィルタ(p)
        if ref:
            m = re.search(r"\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)", ref)
            if m and int(m.group(2)) != 最終データ行:
                w.append(f"オートフィルタ範囲 {ref} がデータ最終行 {最終データ行} と合わない"
                         f"（{os.path.basename(p)}）")
        if 最終行 and 最終行 > 最終データ行:
            w.append(f"データ最終行 {最終データ行} より下に {最終行 - 最終データ行} 行残っている"
                     f"（{os.path.basename(p)}）")
    return w


# --- Excel 出力 ------------------------------------------------------------

FONT = "Arial"
F_TITLE = Font(name=FONT, bold=True, size=13)
F_HEAD  = Font(name=FONT, bold=True, size=10, color="FFFFFF")
F_BOLD  = Font(name=FONT, bold=True, size=10)
F_BODY  = Font(name=FONT, size=10)
F_NOTE  = Font(name=FONT, size=9, italic=True, color="666666")
FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_INFO = PatternFill("solid", fgColor="D9E1F2")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG   = PatternFill("solid", fgColor="FCE4E4")
FILL_OK   = PatternFill("solid", fgColor="E2EFDA")
_thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def 見出し(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c)
        cell.font, cell.fill, cell.border = F_HEAD, FILL_HEAD, BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gl(i)].width = w
    ws.freeze_panes = ws.cell(row + 1, 1)


def 行(ws, r, vals, fill=None, wrap=False):
    for i, v in enumerate(vals, 1):
        cell = ws.cell(r, i, v)
        cell.font, cell.border = F_BODY, BOX
        cell.alignment = Alignment(vertical="top" if wrap else "center", wrap_text=wrap)
        if fill:
            cell.fill = fill


def 出力(results, kintone_data, kyuu_data, meta, out_path):
    wb = openpyxl.Workbook()
    makers = list(results)
    総共通 = sum(len(results[m]["common"]) for m in makers)

    # ---- 1_サマリ ----
    ws = wb.active
    ws.title = "1_サマリ"
    ws["A1"] = "キントーン × 旧システム  セール価格設定リスト 突合結果"
    ws["A1"].font = F_TITLE
    for i, t in enumerate([
            f"作成日: {datetime.date.today():%Y/%m/%d}",
            f"キントーン側: {meta['kintone_dir']}",
            f"旧システム側: {meta['kyuu_dir']}",
            "突合キー: ブランド品番（キントーン C列 / 旧 E列）"], start=2):
        ws.cell(i, 1, t).font = F_NOTE

    r = 7
    ws.cell(r, 1, "■ 件数サマリ").font = F_BOLD
    r += 1
    見出し(ws, r, ["メーカーNo.", "メーカー名", "キントーン行数", "うち在庫あり", "旧システム行数",
                   "共通(突合成立)", "キントーンのみ", "旧システムのみ"],
           [12, 34, 14, 12, 14, 14, 14, 14])
    r += 1
    for m in makers:
        R = results[m]
        行(ws, r, [m, meta["names"].get(m, ""), len(kintone_data[m]), R["kintone_stock"],
                   len(kyuu_data[m]), len(R["common"]), len(R["kintone_only"]), len(R["kyuu_only"])])
        r += 1
    r += 2

    ws.cell(r, 1, "■ 属性項目の一致状況（抽出元に依らない ＝ 不一致は移行ロジックの問題）").font = F_BOLD
    r += 1
    見出し(ws, r, ["比較項目", "キントーン列", "旧システム列", "不一致件数", "判定"],
           [26, 14, 14, 18, 60])
    r += 1
    for label, ki, oi in ATTR:
        n = sum(results[m]["attr_mis"][label] for m in makers)
        ok = n == 0
        行(ws, r, [label, gl(ki), gl(oi), f"{n}/{総共通}",
                   "○ 完全一致" if ok else "△ 要確認 → シート3参照"],
           FILL_OK if ok else FILL_WARN)
        r += 1
    r += 2

    ws.cell(r, 1, "■ 実績項目の一致状況（同じ回どうしのため、本来は一致すべき）").font = F_BOLD
    r += 1
    見出し(ws, r, ["比較項目", "キントーン列", "旧システム列", "不一致件数", "備考"],
           [26, 14, 14, 18, 60])
    r += 1
    for label, ki, oi in PERF:
        n = sum(results[m]["perf_mis"][label] for m in makers)
        行(ws, r, [label, gl(ki), gl(oi), f"{n}/{総共通}",
                   "抽出条件の要調査 ※シート4"], FILL_WARN if n else FILL_OK)
        r += 1

    # ---- 2_項目マッピング ----
    ws = wb.create_sheet("2_項目マッピング")
    ws["A1"] = "項目マッピング（キントーン23列 ⇔ 旧システム45列）"
    ws["A1"].font = F_TITLE
    r = 3
    見出し(ws, r, ["区分", "比較項目", "キントーン列", "キントーン項目名", "旧システム列", "旧システム項目名"],
           [16, 28, 14, 32, 14, 32])
    r += 1
    行(ws, r, ["キー", "ブランド品番", gl(KINTONE_KEYCOL), "ブランド品番", gl(KYUU_KEYCOL), "ブランド品番"], FILL_OK)
    r += 1
    for label, ki, oi in ATTR:
        行(ws, r, ["属性(移行ロジック)", label, gl(ki), KINTONE_NAME[ki], gl(oi), KYUU_NAME[oi]])
        r += 1
    for label, ki, oi in PERF:
        行(ws, r, ["実績(抽出条件)", label, gl(ki), KINTONE_NAME[ki], gl(oi), KYUU_NAME[oi]], FILL_INFO)
        r += 1
    行(ws, r, ["入力欄", "オフ率入力", "N", KINTONE_NAME[14], "O", KYUU_NAME[15]]); r += 1
    行(ws, r, ["入力欄", "税込価格入力", "O", KINTONE_NAME[15], "Q", KYUU_NAME[17]]); r += 2
    ws.cell(r, 1, f"■ 旧システムにあってキントーンに無い項目（{len(ONLY_KYUU)}項目）").font = F_BOLD
    r += 1
    for x in ONLY_KYUU:
        名, 列 = x.split("(")[0], x.split("(")[-1].rstrip(")")
        行(ws, r, ["欠落", 名, "－", "－", 列, 名], FILL_NG)
        r += 1
    r += 1
    ws.cell(r, 1, "■ キントーンにあって旧システムに無い項目").font = F_BOLD
    r += 1
    for x in ONLY_KINTONE:
        名, 列 = x.split("(")[0], x.split("(")[-1].rstrip(")")
        行(ws, r, ["追加", 名, 列, 名, "－", "－"], FILL_WARN)
        r += 1

    # ---- 3_値違い_属性 / 4_値違い_実績 ----
    for title, bucket, fill, note in (
            ("3_値違い_属性", "rows_attr", FILL_NG,
             "共通品番のうち、抽出元に依らない属性項目で値が異なるもの。原則ゼロであるべき。"),
            ("4_値違い_実績", "rows_perf", FILL_WARN,
             "同じ回のリストどうしなので本来は一致するはず。不一致は抽出条件の要調査。在庫から順に追う。")):
        ws = wb.create_sheet(title)
        ws["A1"] = title[2:]
        ws["A1"].font = F_TITLE
        ws["A2"] = note
        ws["A2"].font = F_NOTE
        r = 4
        見出し(ws, r, ["メーカーNo.", "ブランド品番", "品名", "項目", "旧システムの値", "キントーンの値"],
               [12, 40, 44, 22, 24, 24])
        r += 1
        empty = True
        for m in makers:
            for row in results[m][bucket]:
                行(ws, r, [m] + row, fill)
                r += 1
                empty = False
        if empty:
            行(ws, r, ["", "不一致なし", "", "", "", ""], FILL_OK)

    # ---- 5_キントーンのみ ----
    ws = wb.create_sheet("5_キントーンのみ")
    ws["A1"] = "キントーンにのみ存在する品番"
    ws["A1"].font = F_TITLE
    ws["A2"] = "旧システム（基幹システム）は在庫>0 の品番のみ出力するため、在庫なしはここに集まる。"
    ws["A2"].font = F_NOTE
    r = 4
    見出し(ws, r, ["メーカーNo.", "ブランド品番", "品名", "親カテゴリ", "ブランド",
                   "プロパー価格", "在庫", "販売開始日", "推定理由"],
           [12, 40, 44, 16, 20, 14, 10, 14, 34])
    r += 1
    for m in makers:
        K = kintone_data[m]
        for key in results[m]["kintone_only"]:
            v = K[key]
            有 = 在庫あり(v)
            行(ws, r, [m, key, 正規化(v[6]), 正規化(v[0]), 正規化(v[1]), v[7], v[19], 正規化(v[18]),
                       "旧の出力時点では未登録（新規商品）" if 有 else "在庫なし → 旧システムの出力対象外"],
              FILL_WARN if 有 else None)
            r += 1

    # ---- 6_旧システムのみ ----
    ws = wb.create_sheet("6_旧システムのみ")
    ws["A1"] = "旧システムにのみ存在する品番　※移行漏れの可能性"
    ws["A1"].font = F_TITLE
    r = 3
    見出し(ws, r, ["メーカーNo.", "ブランド品番", "品名", "親カテゴリー", "ブランド",
                   "元上代", "在庫", "販売可能数", "SKU数"],
           [12, 40, 44, 16, 20, 14, 12, 12, 10])
    r += 1
    empty = True
    for m in makers:
        O = kyuu_data[m]
        for key in results[m]["kyuu_only"]:
            v = O[key]
            行(ws, r, [m, key, 正規化(v[8]), 正規化(v[2]), 正規化(v[3]), v[10], v[25], v[24], v[9]], FILL_NG)
            r += 1
            empty = False
    if empty:
        行(ws, r, ["", "該当なし", "", "", "", "", "", "", ""], FILL_OK)

    for s in wb.worksheets:
        s.sheet_view.showGridLines = False
    wb.save(out_path)


# --- エントリポイント ------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description="キントーンと旧システムのセール価格設定リストを突合する")
    ap.add_argument("--kintone", default=KINTONE_LIST, help="キントーン出力の xlsx を置いたフォルダ")
    ap.add_argument("--旧", dest="kyuu", default=KYUU_LIST, help="旧システムが出力した送信フォルダ")
    ap.add_argument("--out", default=None, help="出力する突合結果 xlsx（既定: キントーン側フォルダ）")
    ap.add_argument("--メーカー", dest="makers", nargs="*", default=None, help="対象メーカーNo.（既定: 全件）")
    a = ap.parse_args(argv)

    kintone_files, kyuu_files = 一覧(a.kintone), 一覧(a.kyuu)
    targets = [m for m in sorted(kintone_files, key=int) if m in kyuu_files]
    if a.makers:
        want = {str(int(x)) for x in a.makers}
        targets = [m for m in targets if m in want]

    片方だけ = sorted(set(kintone_files) ^ set(kyuu_files), key=int)
    if not targets:
        print("突合できるメーカーがありません。")
        print(f"  キントーン: {a.kintone} … {len(kintone_files)}メーカー")
        print(f"  旧システム: {a.kyuu} … {len(kyuu_files)}メーカー")
        return 1

    results, kintone_data, kyuu_data, names = {}, {}, {}, {}
    警告 = {}
    for m in targets:
        K, dupK = 読む(kintone_files[m], KINTONE_KEYCOL, KINTONE_NCOL)
        O, dupO = 読む(kyuu_files[m], KYUU_KEYCOL, KYUU_NCOL)
        kintone_data[m], kyuu_data[m] = K, O
        # メーカー名は旧システム側の B列が正（キントーン側は持っていない）
        names[m] = 正規化(next(iter(O.values()))[1]) if O else ""
        results[m] = 突合(K, O)
        R = results[m]
        print(f"メーカー{m:>3} {names[m]}")
        print(f"    キントーン {len(K):>5}行（在庫あり {R['kintone_stock']}）"
              f" / 旧 {len(O):>5}行 / 共通 {len(R['common']):>5}"
              f" / kのみ {len(R['kintone_only']):>5} / 旧のみ {len(R['kyuu_only'])}")
        print(f"    属性の不一致 {sum(R['attr_mis'].values())} 件"
              f" / 実績の不一致 {sum(R['perf_mis'].values())} 件")
        for key, fn in dupK + dupO:
            print(f"    ※ 品番の重複: {key}（{fn}）")
        警告[m] = 健全性(K, O, R, kintone_files[m])
        for msg in 警告[m]:
            print(f"    [警告] {msg}")

    out = a.out or os.path.join(a.kintone,
                                f"{datetime.date.today():%y%m%d}_kintone_旧システム_突合結果.xlsx")
    出力(results, kintone_data, kyuu_data,
         dict(kintone_dir=a.kintone, kyuu_dir=a.kyuu, names=names), out)

    総属性 = sum(sum(results[m]["attr_mis"].values()) for m in targets)
    print(f"\n出力: {out}")
    print(f"対象 {len(targets)} メーカー / 属性の不一致 合計 {総属性} 件")

    総警告 = sum(len(v) for v in 警告.values())
    if 総警告:
        print(f"\n[警告 {総警告}件] 一致件数だけでは気づけない異常がある")
        for m in targets:
            for msg in 警告.get(m, []):
                print(f"  メーカー{m:>3}: {msg}")
    else:
        print("健全性チェック: 異常なし")
    if 片方だけ:
        頭 = ", ".join(片方だけ[:15])
        続き = f" …ほか{len(片方だけ) - 15}メーカー" if len(片方だけ) > 15 else ""
        print(f"※ 片方にしか無いメーカー（{len(片方だけ)}件）: {頭}{続き}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
