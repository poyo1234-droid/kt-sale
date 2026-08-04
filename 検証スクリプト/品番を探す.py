"""ブランド品番が、どの TXT ファイルの何行目・どのシートにあるかを調べる。

  python 品番を探す.py GLO-GTP9582
  python 品番を探す.py NOB-604/605

親カテゴリ単位で見たいときは --親カテ を使う。

  python 品番を探す.py --親カテ "EVER METAL"
"""
import os, sys, glob, collections
import openpyxl
from _共通 import REPO, TXT_TANTOUSHA, TXT_KENSHOU, read_txt

BOOK = os.path.join(REPO, "_テスト用", "sale処理-v53_TXT出力.xlsm")


def 品番で探す(pn):
    print(f"品番: {pn}\n=== TXT 内の所在 ===")
    for label, d in (("担当者", TXT_TANTOUSHA), ("検証  ", TXT_KENSHOU)):
        if not os.path.isdir(d):
            continue
        for p in sorted(glob.glob(os.path.join(d, "*.txt"))):
            for i, line in enumerate(read_txt(p), start=1):
                if line.split("\t")[0] == pn:
                    print(f"  [{label}] {os.path.basename(p)} {i}行目: {line}")

    if not os.path.exists(BOOK):
        return
    print("\n=== シート内の所在 ===")
    wb = openpyxl.load_workbook(BOOK, data_only=True, read_only=True)
    for name in wb.sheetnames:
        hdr = None
        for i, r in enumerate(wb[name].iter_rows(values_only=True), start=1):
            if i == 1:
                hdr = [str(c) if c is not None else "" for c in r]
                continue
            if any(str(c) == pn for c in r if c is not None):
                cols = {hdr[j] if j < len(hdr) else f"col{j}": r[j]
                        for j in range(min(len(r), 8))}
                print(f"  「{name}」{i}行目: {cols}")
                break


def 親カテで探す(cat, sheet="通年-記入以外前回と同じ"):
    wb = openpyxl.load_workbook(BOOK, data_only=True, read_only=True)
    ws = wb[sheet]
    pns, pos = set(), []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r[2] and r[1] == cat:      # B列=親カテゴリ / C列=ブランド品番
            pns.add(str(r[2]))
            pos.append(i)
    print(f"親カテゴリ「{cat}」: {len(pns)}品番  シート行 {min(pos)}〜{max(pos)}"
          f"  連続ブロック={pos == list(range(min(pos), max(pos)+1))}")

    for label, d in (("担当者", TXT_TANTOUSHA), ("検証", TXT_KENSHOU)):
        if not os.path.isdir(d):
            continue
        hit = collections.Counter()
        for p in sorted(glob.glob(os.path.join(d, sheet + "_*.txt"))):
            for line in read_txt(p):
                if line.split("\t")[0] in pns:
                    hit[os.path.basename(p)] += 1
        print(f"  [{label}] 計 {sum(hit.values())}件")
        for f, n in sorted(hit.items()):
            print(f"     {f}  {n}件")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    if sys.argv[1] == "--親カテ":
        親カテで探す(sys.argv[2])
    else:
        品番で探す(sys.argv[1])
